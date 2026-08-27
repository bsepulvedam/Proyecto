import logging
import re
import unicodedata
from zipfile import BadZipFile
from collections import Counter
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.empresa import Empresa
from app.models.producto import Producto
from app.models.unidad_medida import UnidadMedida


SOURCE_FILENAME = "Control_Inventario_Bodega_Boliklor_ALM_DEFI.xlsx"
MASTER_SHEET = "Maestro de materiales"
STOCK_SHEETS = {"Stock Boliklor": "BOLIKLOR", "Stock ALM": "ALM"}
UNIT_ALIASES = {"UNIDAD": "UN", "UND": "UN"}
TYPO_MARKERS = {
    "PEGAMETO": "Posible typo: PEGAMETO / PEGAMENTO.",
    "INSTALCION": "Posible typo: INSTALCION / INSTALACION.",
}
logger = logging.getLogger(__name__)


class ProductImportError(RuntimeError):
    pass


class ProductImportExecutionError(RuntimeError):
    pass


@dataclass
class StockReference:
    sheet: str
    company_code: str
    sku: str
    name: str
    unit: str
    minimum_stock: object


@dataclass
class ProductoImportRow:
    fila_excel: int
    sku: str
    empresa: str | None
    nombre: str
    descripcion: str | None
    unidad_stock: str | None
    unidad_contenido: str | None
    factor_conversion: Decimal | None
    unidad_costo: str | None
    stock_minimo: Decimal | None
    tipo: str
    familia: str
    catalog_status: str = "NUEVO"
    errores: list[str] = field(default_factory=list)
    advertencias: list[str] = field(default_factory=list)

    @property
    def listo_para_importar(self) -> bool:
        return not self.errores and self.catalog_status != "CONFLICTO"

    @property
    def validation_status(self) -> str:
        if self.errores:
            return "ERROR"
        if self.advertencias:
            return "ADVERTENCIA"
        return "VALIDO"


@dataclass
class ProductImportReport:
    source_file: str
    sheets_used: list[str]
    rows: list[ProductoImportRow]
    unidades_detectadas: list[str]
    unidades_no_registradas: list[str]
    duplicados: list[str]
    conflictos_maestro_stock: list[str]
    productos_stock_sin_maestro: list[str]

    @property
    def total(self) -> int:
        return len(self.rows)

    @property
    def total_boliklor(self) -> int:
        return sum(row.empresa == "BOLIKLOR" for row in self.rows)

    @property
    def total_alm(self) -> int:
        return sum(row.empresa == "ALM" for row in self.rows)

    @property
    def validos(self) -> int:
        return sum(row.validation_status == "VALIDO" for row in self.rows)

    @property
    def con_advertencias(self) -> int:
        return sum(row.validation_status == "ADVERTENCIA" for row in self.rows)

    @property
    def con_errores(self) -> int:
        return sum(row.validation_status == "ERROR" for row in self.rows)


@dataclass
class ProductImportResult:
    analizados: int
    importados: int
    existentes: int
    advertencias_pendientes: int
    errores_pendientes: int
    fallidos: list[str]

    @property
    def total_fallidos(self) -> int:
        return len(self.fallidos)


def default_source_path() -> Path:
    return Path.home() / "Desktop" / SOURCE_FILENAME


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def normalize_sku(value: object) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def normalize_unit(value: object) -> str:
    unit = normalize_text(value)
    return UNIT_ALIASES.get(unit, unit)


def parse_decimal(value: object) -> Decimal | None:
    if value is None or normalize_text(value) == "":
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def company_from_sku(sku: str) -> str | None:
    if sku.startswith("BOL-"):
        return "BOLIKLOR"
    if sku.startswith("ALM-"):
        return "ALM"
    return None


def unit_mapping(stock_unit: str, factor: Decimal | None) -> tuple[str | None, str | None]:
    if not stock_unit:
        return None, None
    if factor is not None and factor > 1 and stock_unit == "SACO":
        return "KG", "KG"
    return None, stock_unit


def _normalized_sheet_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _find_sheet(workbook, expected: str):
    expected_name = _normalized_sheet_name(expected)
    for sheet_name in workbook.sheetnames:
        if _normalized_sheet_name(sheet_name) == expected_name:
            return workbook[sheet_name]
    raise ProductImportError(f'No se encontró la hoja requerida "{expected}".')


def _validate_headers(sheet, row_number: int, expected: list[str]) -> None:
    values = next(
        sheet.iter_rows(
            min_row=row_number,
            max_row=row_number,
            min_col=1,
            max_col=len(expected),
            values_only=True,
        )
    )
    normalized = [normalize_text(value) for value in values]
    if normalized != expected:
        raise ProductImportError(
            f"Los encabezados de la hoja {sheet.title} no coinciden con el formato esperado."
        )


def _stock_references(workbook) -> tuple[dict[str, list[StockReference]], list[str]]:
    references: dict[str, list[StockReference]] = {}
    used_sheets: list[str] = []
    for expected_sheet, company_code in STOCK_SHEETS.items():
        sheet = _find_sheet(workbook, expected_sheet)
        _validate_headers(
            sheet,
            4,
            ["SKU", "ITEM", "UNIDAD", "TIPO", "FAMILIA", "STOCK MIN"],
        )
        used_sheets.append(sheet.title)
        for excel_row, values in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
            sku = normalize_sku(values[0])
            if not sku:
                continue
            references.setdefault(sku, []).append(
                StockReference(
                    sheet=sheet.title,
                    company_code=company_code,
                    sku=sku,
                    name=normalize_text(values[1]),
                    unit=normalize_unit(values[2]),
                    minimum_stock=values[5],
                )
            )
    return references, used_sheets


def _existing_product_conflicts(
    row: ProductoImportRow,
    existing: Producto,
) -> list[str]:
    conflicts: list[str] = []
    if existing.empresa.codigo != row.empresa:
        conflicts.append("La empresa difiere del producto existente.")
    if normalize_text(existing.nombre) != row.nombre:
        conflicts.append("El nombre difiere del producto existente.")
    if existing.unidad_stock.codigo != row.unidad_stock:
        conflicts.append("La unidad de stock difiere del producto existente.")
    existing_content = existing.unidad_contenido.codigo if existing.unidad_contenido else None
    if existing_content != row.unidad_contenido:
        conflicts.append("La unidad de contenido difiere del producto existente.")
    if existing.factor_conversion != row.factor_conversion:
        conflicts.append("El factor de conversión difiere del producto existente.")
    if existing.unidad_costo.codigo != row.unidad_costo:
        conflicts.append("La unidad de costo difiere del producto existente.")
    if existing.stock_minimo != row.stock_minimo:
        conflicts.append("El stock mínimo difiere del producto existente.")
    if normalize_text(existing.tipo) != row.tipo:
        conflicts.append("El tipo difiere del producto existente.")
    if normalize_text(existing.familia) != row.familia:
        conflicts.append("La familia difiere del producto existente.")
    return conflicts


def analyze_product_workbook(
    source_path: Path,
    companies: list[Empresa],
    units: list[UnidadMedida],
    existing_products: list[Producto],
) -> ProductImportReport:
    if not source_path.is_file():
        raise ProductImportError(f"No se encontró el archivo fuente {source_path.name}.")

    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise ProductImportError("El archivo fuente no es un libro Excel válido.") from exc
    try:
        master = _find_sheet(workbook, MASTER_SHEET)
        _validate_headers(
            master,
            1,
            ["SKU", "ITEM", "UNIDAD", "TIPO", "FAMILIA", "UNIDAD"],
        )
        stock_by_sku, stock_sheets = _stock_references(workbook)
        company_codes = {company.codigo for company in companies}
        unit_codes = {unit.codigo for unit in units}
        existing_by_sku = {normalize_sku(product.sku): product for product in existing_products}
        raw_rows: list[tuple[int, tuple[object, ...]]] = []
        for excel_row, values in enumerate(master.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
            if any(value is not None and str(value).strip() for value in values):
                raw_rows.append((excel_row, values))

        sku_counts = Counter(normalize_sku(values[0]) for _, values in raw_rows if normalize_sku(values[0]))
        duplicates = sorted(sku for sku, count in sku_counts.items() if count > 1)
        rows: list[ProductoImportRow] = []
        units_detected: set[str] = set()
        unknown_units: set[str] = set()
        master_stock_conflicts: list[str] = []

        for excel_row, values in raw_rows:
            sku = normalize_sku(values[0])
            name = normalize_text(values[1])
            stock_unit = normalize_unit(values[2])
            product_type = normalize_text(values[3])
            family = normalize_text(values[4])
            factor = parse_decimal(values[5])
            company_code = company_from_sku(sku)
            content_unit, cost_unit = unit_mapping(stock_unit, factor)
            references = stock_by_sku.get(sku, [])
            minimum_stock = parse_decimal(references[0].minimum_stock) if len(references) == 1 else None

            row = ProductoImportRow(
                fila_excel=excel_row,
                sku=sku,
                empresa=company_code,
                nombre=name,
                descripcion=None,
                unidad_stock=stock_unit or None,
                unidad_contenido=content_unit,
                factor_conversion=factor,
                unidad_costo=cost_unit,
                stock_minimo=minimum_stock,
                tipo=product_type,
                familia=family,
            )

            if not sku:
                row.errores.append("SKU vacío.")
            if sku in duplicates:
                row.errores.append("SKU duplicado en Maestro de materiales.")
            if not name:
                row.errores.append("Nombre vacío.")
            if company_code is None:
                row.errores.append("Empresa no identificada desde el SKU.")
            elif company_code not in company_codes:
                row.errores.append(f"Empresa {company_code} no registrada.")
            if not product_type:
                row.errores.append("Tipo vacío.")
            if not family:
                row.errores.append("Familia vacía.")
            if factor is None:
                row.errores.append("Factor de conversión vacío o no numérico.")
            elif factor <= 0:
                row.errores.append("Factor de conversión debe ser mayor que 0.")
            if minimum_stock is None:
                if not references:
                    row.errores.append("Producto de Maestro sin registro en la hoja de stock correspondiente.")
                else:
                    row.errores.append("Stock mínimo vacío o no numérico.")
            elif minimum_stock < 0:
                row.errores.append("Stock mínimo no puede ser negativo.")

            for unit_label, unit_code in (
                ("stock", row.unidad_stock),
                ("contenido", row.unidad_contenido),
                ("costo", row.unidad_costo),
            ):
                if not unit_code:
                    if unit_label == "stock":
                        row.errores.append("Unidad de stock vacía.")
                    continue
                units_detected.add(unit_code)
                if unit_code not in unit_codes:
                    unknown_units.add(unit_code)
                    row.errores.append(f"Unidad de {unit_label} no registrada: {unit_code}.")

            if len(references) > 1:
                row.errores.append("SKU repetido en hojas de stock.")
            elif references:
                reference = references[0]
                if reference.company_code != company_code:
                    message = f"{sku}: empresa {company_code} no coincide con {reference.sheet}."
                    row.errores.append(message)
                    master_stock_conflicts.append(message)
                if reference.unit != stock_unit:
                    message = f"{sku}: unidad Maestro {stock_unit or 'VACÍA'} / Stock {reference.unit or 'VACÍA'}."
                    row.errores.append("Unidad inconsistente entre Maestro y Stock.")
                    master_stock_conflicts.append(message)
                if reference.name != name:
                    row.advertencias.append(
                        f"Nombre distinto en Stock: {reference.name}. Revisar sin corregir automáticamente."
                    )

            text_for_typos = f"{name} {product_type} {family}"
            for marker, warning in TYPO_MARKERS.items():
                if marker in text_for_typos:
                    row.advertencias.append(warning)

            existing = existing_by_sku.get(sku)
            if existing is not None:
                conflicts = _existing_product_conflicts(row, existing)
                if conflicts:
                    row.catalog_status = "CONFLICTO"
                    row.errores.extend(conflicts)
                else:
                    row.catalog_status = "EXISTENTE"
            rows.append(row)

        master_skus = {row.sku for row in rows if row.sku}
        stock_without_master = sorted(set(stock_by_sku) - master_skus)
        return ProductImportReport(
            source_file=source_path.name,
            sheets_used=[master.title, *stock_sheets],
            rows=rows,
            unidades_detectadas=sorted(units_detected),
            unidades_no_registradas=sorted(unknown_units),
            duplicados=duplicates,
            conflictos_maestro_stock=sorted(set(master_stock_conflicts)),
            productos_stock_sin_maestro=stock_without_master,
        )
    finally:
        workbook.close()


def import_valid_products(db: Session, report: ProductImportReport) -> ProductImportResult:
    """Importa candidatos sin errores/advertencias en un único commit.

    Las filas cuyo catálogo no pueda resolverse se excluyen con un motivo exacto.
    Un fallo inesperado de infraestructura revierte todas las inserciones preparadas.
    """

    valid_rows = [row for row in report.rows if row.validation_status == "VALIDO"]
    logger.info(
        "Inicio importación de productos: analizados=%s validos=%s",
        report.total,
        len(valid_rows),
    )
    try:
        companies = {company.codigo: company for company in db.scalars(select(Empresa)).all()}
        units = {unit.codigo: unit for unit in db.scalars(select(UnidadMedida)).all()}
        existing_skus = set(db.scalars(select(Producto.sku)).all())

        imported = 0
        existing = 0
        failures: list[str] = []
        for row in valid_rows:
            if row.sku in existing_skus:
                existing += 1
                continue

            company = companies.get(row.empresa or "")
            stock_unit = units.get(row.unidad_stock or "")
            content_unit = units.get(row.unidad_contenido) if row.unidad_contenido else None
            cost_unit = units.get(row.unidad_costo or "")
            unresolved: list[str] = []
            if company is None:
                unresolved.append(f"empresa {row.empresa or 'vacía'}")
            if stock_unit is None:
                unresolved.append(f"unidad stock {row.unidad_stock or 'vacía'}")
            if row.unidad_contenido and content_unit is None:
                unresolved.append(f"unidad contenido {row.unidad_contenido}")
            if cost_unit is None:
                unresolved.append(f"unidad costo {row.unidad_costo or 'vacía'}")
            if unresolved:
                message = f"{row.sku}: no se pudo resolver " + ", ".join(unresolved)
                failures.append(message)
                logger.warning(message)
                continue

            db.add(
                Producto(
                    empresa_id=company.id,
                    sku=row.sku,
                    nombre=row.nombre,
                    descripcion=row.descripcion,
                    unidad_stock_id=stock_unit.id,
                    unidad_contenido_id=content_unit.id if content_unit else None,
                    factor_conversion=row.factor_conversion,
                    unidad_costo_id=cost_unit.id,
                    stock_minimo=row.stock_minimo,
                    tipo=row.tipo,
                    familia=row.familia,
                    activo=True,
                )
            )
            existing_skus.add(row.sku)
            imported += 1

        db.commit()
        result = ProductImportResult(
            analizados=report.total,
            importados=imported,
            existentes=existing,
            advertencias_pendientes=report.con_advertencias,
            errores_pendientes=report.con_errores,
            fallidos=failures,
        )
        logger.info(
            "Fin importación de productos: importados=%s existentes=%s fallidos=%s",
            result.importados,
            result.existentes,
            result.total_fallidos,
        )
        return result
    except Exception as exc:
        db.rollback()
        logger.exception("Error técnico durante la importación de productos")
        raise ProductImportExecutionError(
            "No fue posible completar la importación de productos."
        ) from exc
