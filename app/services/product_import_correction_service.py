from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal

from openpyxl import load_workbook

from app.services.product_import_service import (
    ProductImportError, _find_sheet, company_from_sku, normalize_sku,
    normalize_text, parse_decimal,
)

PRESENTATION_ALIASES = {"UN": "UNIDAD", "UND": "UNIDAD", "UNIDAD": "UNIDAD"}
STOCK_SHEETS = ("Stock Boliklor", "Stock ALM")


def normalize_presentation(value) -> str:
    presentation = normalize_text(value)
    return PRESENTATION_ALIASES.get(presentation, presentation)


def format_quantity(value: Decimal | None) -> str:
    if value is None:
        return "—"
    return str(int(value)) if value == value.to_integral_value() else format(value.normalize(), "f")


@dataclass(frozen=True)
class ReceptionEvidence:
    quantity: Decimal
    unit_cost: Decimal
    factor: Decimal
    purchase_price: Decimal
    presentation: str

    @property
    def formula_matches(self) -> bool:
        return self.unit_cost * self.quantity * self.factor == self.purchase_price


@dataclass(frozen=True)
class ProductCorrectionRow:
    sku: str
    empresa: str | None
    nombre: str
    unidad_presentacion: str | None
    stock_actual: Decimal
    stock_minimo: Decimal | None
    factor_conversion: Decimal | None
    unidad_contenido: str | None
    unidad_costo: str | None
    tipo: str
    familia: str
    estado_stock: str | None
    estado_validacion: str
    observaciones: list[str] = field(default_factory=list)

    @property
    def stock_actual_display(self) -> str:
        return format_quantity(self.stock_actual)

    @property
    def stock_minimo_display(self) -> str:
        return format_quantity(self.stock_minimo)

    @property
    def factor_display(self) -> str:
        return format_quantity(self.factor_conversion)


@dataclass(frozen=True)
class ProductCorrectionReport:
    source_file: str
    rows: list[ProductCorrectionRow]
    unidades_detectadas: list[str]
    unidades_no_reconocidas: list[str]

    @property
    def total(self) -> int:
        return len(self.rows)

    @property
    def listos(self) -> int:
        return sum(row.estado_validacion == "LISTO" for row in self.rows)

    @property
    def advertencias(self) -> int:
        return sum(row.estado_validacion == "ADVERTENCIA" for row in self.rows)

    @property
    def errores(self) -> int:
        return sum(row.estado_validacion == "ERROR" for row in self.rows)


def _sum_movements(sheet, quantity_column: int) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for values in sheet.iter_rows(min_row=2, max_col=quantity_column, values_only=True):
        sku = normalize_sku(values[0])
        quantity = parse_decimal(values[quantity_column - 1])
        if sku and quantity is not None:
            totals[sku] += quantity
    return totals


def analyze_legacy_stock(workbook) -> dict[str, Decimal]:
    receptions = _sum_movements(_find_sheet(workbook, "Recepción"), 4)
    dispatches = _sum_movements(_find_sheet(workbook, "Despacho"), 6)
    returns = _sum_movements(_find_sheet(workbook, "Devoluciones"), 4)
    skus = set(receptions) | set(dispatches) | set(returns)
    return {
        sku: receptions.get(sku, Decimal("0"))
        - dispatches.get(sku, Decimal("0"))
        + returns.get(sku, Decimal("0"))
        for sku in skus
    }


def _stock_references(workbook):
    references = {}
    for sheet_name in STOCK_SHEETS:
        sheet = _find_sheet(workbook, sheet_name)
        for values in sheet.iter_rows(min_row=5, max_col=6, values_only=True):
            sku = normalize_sku(values[0])
            if sku:
                references[sku] = (normalize_presentation(values[2]), parse_decimal(values[5]))
    return references


def _reception_evidence(workbook) -> dict[str, list[ReceptionEvidence]]:
    evidence = defaultdict(list)
    sheet = _find_sheet(workbook, "Recepción")
    for values in sheet.iter_rows(min_row=2, max_col=10, values_only=True):
        sku = normalize_sku(values[0])
        quantity, unit_cost = parse_decimal(values[3]), parse_decimal(values[4])
        factor, purchase_price = parse_decimal(values[5]), parse_decimal(values[6])
        if sku and None not in (quantity, unit_cost, factor, purchase_price):
            evidence[sku].append(ReceptionEvidence(
                quantity=quantity, unit_cost=unit_cost, factor=factor,
                purchase_price=purchase_price,
                presentation=normalize_presentation(values[9]),
            ))
    return evidence


def analyze_product_corrections(source_path, products, units):
    """Genera la vista legacy de solo lectura; nunca modifica PostgreSQL."""
    if not source_path.is_file():
        raise ProductImportError(f"No se encontró el archivo fuente {source_path.name}.")
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        master = _find_sheet(workbook, "Maestro de materiales")
        master_by_sku = {
            normalize_sku(values[0]): values
            for values in master.iter_rows(min_row=2, max_col=6, values_only=True)
            if normalize_sku(values[0])
        }
        stock_references = _stock_references(workbook)
        legacy_stock = analyze_legacy_stock(workbook)
        reception = _reception_evidence(workbook)
        registered = {normalize_presentation(unit.codigo) for unit in units}
        detected, unknown, rows = set(), set(), []

        for product in products:
            sku = normalize_sku(product.sku)
            values = master_by_sku.get(sku)
            errors, warnings = [], []
            company = company_from_sku(sku)
            presentation = factor = minimum = None
            name, product_type, family = normalize_text(product.nombre), "", ""
            if values is None:
                errors.append("SKU no encontrado en Maestro de materiales.")
            else:
                name = normalize_text(values[1])
                presentation = normalize_presentation(values[2]) or None
                product_type, family = normalize_text(values[3]), normalize_text(values[4])
                factor = parse_decimal(values[5])
            stock_reference = stock_references.get(sku)
            if stock_reference:
                stock_presentation, minimum = stock_reference
                if presentation and stock_presentation != presentation:
                    errors.append("Presentación inconsistente entre Maestro y hoja Stock.")
            else:
                errors.append("SKU no encontrado en la hoja Stock correspondiente.")
            if not company:
                errors.append("Empresa no identificable desde el SKU.")
            if not presentation:
                errors.append("Presentación vacía en el Maestro.")
            elif presentation not in registered:
                errors.append(f"Presentación no registrada: {presentation}.")
                unknown.add(presentation)
            else:
                detected.add(presentation)
            if factor is None or factor <= 0:
                errors.append("Factor vacío, no numérico o no positivo.")
            if minimum is None or minimum < 0:
                errors.append("Stock mínimo vacío, no numérico o negativo.")

            content_unit = None
            if presentation == "SACO" and factor and factor > 1 and "PINTURA TERMOPLASTICA" in name:
                content_unit = "KG"
                detected.add("KG")
            matching_evidence = [item for item in reception.get(sku, []) if item.formula_matches and item.factor == factor]
            cost_unit = None
            if matching_evidence:
                cost_unit = "KG" if content_unit == "KG" else presentation
            elif presentation == "KG" and factor == 1:
                cost_unit = "KG"
            else:
                warnings.append("Unidad costo no confirmada por evidencia suficiente en Recepción.")
            if factor and factor > 1 and content_unit is None:
                warnings.append("Unidad contenido no confirmada para esta conversión.")

            current_stock = legacy_stock.get(sku, Decimal("0"))
            stock_state = "EN STOCK" if current_stock > 0 else "SIN STOCK"
            validation = "ERROR" if errors else "ADVERTENCIA" if warnings else "LISTO"
            rows.append(ProductCorrectionRow(
                sku=sku, empresa=company, nombre=name,
                unidad_presentacion=presentation, stock_actual=current_stock,
                stock_minimo=minimum, factor_conversion=factor,
                unidad_contenido=content_unit, unidad_costo=cost_unit,
                tipo=product_type, familia=family, estado_stock=stock_state,
                estado_validacion=validation, observaciones=[*errors, *warnings],
            ))
        return ProductCorrectionReport(
            source_file=source_path.name, rows=rows,
            unidades_detectadas=sorted(detected),
            unidades_no_reconocidas=sorted(unknown),
        )
    finally:
        workbook.close()
