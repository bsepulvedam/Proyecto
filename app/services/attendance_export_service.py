from io import BytesIO
import re

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.services.attendance_supervision_service import (
    SupervisionPeriod,
    WorkerSupervision,
    get_worker_supervision,
    iter_worker_supervision_summaries,
)


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MISSING_RATE_LABEL = "Sin tarifa configurada"
FORMULA_PREFIXES = ("=", "+", "-", "@")
ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


class AttendanceExportError(ValueError):
    pass


def safe_excel_text(value: object) -> object:
    if isinstance(value, str):
        cleaned = ILLEGAL_XML_CHARACTERS.sub(" ", value)
        if cleaned.lstrip().startswith(FORMULA_PREFIXES):
            return f"'{cleaned}"
        return cleaned
    return value


def _append_safe(worksheet, values) -> None:
    worksheet.append(tuple(safe_excel_text(value) for value in values))


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_combined_attendance_xlsx(
    db: Session,
    period: SupervisionPeriod,
    *,
    query: str | None = None,
) -> bytes:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Resumen")
    _append_safe(
        sheet,
        (
            "Trabajador",
            "Código",
            "Días trabajados",
            "Jornadas pagables",
            "Dobles turnos",
            "Incidencias",
            "Total provisional",
        ),
    )
    for item in iter_worker_supervision_summaries(db, period, query=query):
        projection = item.projection
        _append_safe(
            sheet,
            (
                f"{item.worker.nombres} {item.worker.apellidos}".strip(),
                item.worker.codigo_interno or "",
                projection.completed_worked_days,
                projection.payable_shifts,
                projection.double_shift_days,
                projection.incident_count,
                (
                    projection.provisional_total_clp
                    if projection.provisional_total_clp is not None
                    else MISSING_RATE_LABEL
                ),
            ),
        )
    return _workbook_bytes(workbook)


def _session_times(day, attribute: str) -> str:
    values: list[str] = []
    for session in day.sessions:
        value = getattr(session.projection, attribute)
        if value is not None:
            values.append(
                f"{session.projection.shift_code}: {value.strftime('%d/%m/%Y %H:%M')}"
            )
    return " | ".join(values)


def build_individual_attendance_xlsx(supervision: WorkerSupervision) -> bytes:
    workbook = Workbook(write_only=True)
    summary = workbook.create_sheet("Resumen")
    worker_name = (
        f"{supervision.worker.nombres} {supervision.worker.apellidos}".strip()
    )
    projection = supervision.projection
    summary_rows = (
        ("Trabajador", worker_name),
        ("Período", f"{supervision.period.start.isoformat()} al {supervision.period.end.isoformat()}"),
        ("Días trabajados", projection.completed_worked_days),
        ("Jornadas pagables", projection.payable_shifts),
        ("Dobles turnos", projection.double_shift_days),
        ("Incidencias", projection.incident_count),
        (
            "Total provisional",
            projection.provisional_total_clp
            if projection.provisional_total_clp is not None
            else MISSING_RATE_LABEL,
        ),
    )
    for row in summary_rows:
        _append_safe(summary, row)

    detail = workbook.create_sheet("Detalle diario")
    _append_safe(
        detail,
        (
            "Fecha",
            "Turno",
            "Estado",
            "Entrada",
            "Salida",
            "Jornada pagable",
            "Tarifa efectiva",
            "Origen tarifa",
            "Total provisional",
            "Incidencias",
        ),
    )
    for day in supervision.days:
        day_projection = day.projection
        rate = day_projection.effective_rate
        _append_safe(
            detail,
            (
                day_projection.operational_date,
                " + ".join(day_projection.payable_shift_codes),
                " · ".join(day.labels),
                _session_times(day, "entry_time"),
                _session_times(day, "exit_time"),
                day_projection.payable_shifts,
                rate.amount_clp if rate is not None else MISSING_RATE_LABEL,
                rate.source if rate is not None else MISSING_RATE_LABEL,
                (
                    day_projection.provisional_total_clp
                    if day_projection.provisional_total_clp is not None
                    else MISSING_RATE_LABEL
                ),
                day_projection.incident_count,
            ),
        )
    return _workbook_bytes(workbook)


def export_worker_attendance(
    db: Session,
    worker_id: int,
    period: SupervisionPeriod,
) -> bytes | None:
    supervision = get_worker_supervision(db, worker_id, period)
    if supervision is None:
        return None
    return build_individual_attendance_xlsx(supervision)


def combined_export_filename(period: SupervisionPeriod) -> str:
    return f"asistencia_{period.start.isoformat()}_{period.end.isoformat()}.xlsx"


def individual_export_filename(worker_id: int, period: SupervisionPeriod) -> str:
    if isinstance(worker_id, bool) or worker_id <= 0:
        raise AttendanceExportError("El trabajador no es válido.")
    return (
        f"asistencia_trabajador_{worker_id}_{period.start.isoformat()}_"
        f"{period.end.isoformat()}.xlsx"
    )
