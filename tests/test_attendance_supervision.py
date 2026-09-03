import os
import re
import unittest
from datetime import date, datetime, time, timezone
from decimal import Decimal
from html import unescape
from io import BytesIO

from openpyxl import load_workbook

from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.time import app_timezone
from app.database.base import Base
from app.database.session import get_db
from app.main import app
from app.models.attendance import (
    EvaluacionGeograficaMarcaje,
    IncidenciaAsistencia,
    LugarTrabajo,
    MarcajeAsistencia,
    SesionTrabajo,
    TarifaProvisionalAsistencia,
    Turno,
)
from app.models.empresa import Empresa
from app.models.identity import Rol, Trabajador
from app.schemas.identity import UserCreate
from app.services.attendance_supervision_service import (
    AttendanceSupervisionError,
    get_worker_day_supervision,
    list_worker_supervision,
    supervision_period,
)
from app.services.auth_service import create_user
from tests.test_identity_auth import ASGIClient


class AttendanceSupervisionTests(unittest.TestCase):
    def setUp(self) -> None:
        names = ("APP_ENV", "AUTH_ENFORCED", "SESSION_SECRET", "COOKIE_SECURE", "APP_TIMEZONE")
        self.previous = {name: os.environ.get(name) for name in names}
        os.environ.update(
            {
                "APP_ENV": "test",
                "AUTH_ENFORCED": "true",
                "SESSION_SECRET": "test-secret-only-not-production",
                "COOKIE_SECURE": "false",
                "APP_TIMEZONE": "America/Santiago",
            }
        )
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        with self.Session() as db:
            db.add_all(
                [
                    Rol(codigo="ADMIN", nombre="Administrador"),
                    Rol(codigo="JEFATURA", nombre="Jefatura"),
                    Rol(codigo="TRABAJADOR", nombre="Trabajador"),
                    Empresa(codigo="SUP", nombre="Supervisión"),
                    Turno(codigo="DIURNO", nombre="Diurno"),
                    Turno(codigo="NOCTURNO", nombre="Nocturno"),
                    LugarTrabajo(
                        nombre="Comuna histórica",
                        tipo="TERRENO",
                        comuna="Colina",
                        tipo_geocerca="COMUNA",
                        codigo_comuna="13301",
                        latitud=Decimal("-33.2"),
                        longitud=Decimal("-70.6"),
                        prioridad_geocerca=10,
                    ),
                ]
            )
            db.commit()
            self.users = {
                "admin": create_user(db, UserCreate(username="admin-supervision", password="Clave-Admin-Supervision-123"), "ADMIN"),
                "jefatura": create_user(db, UserCreate(username="jefatura-supervision", password="Clave-Jefatura-Supervision-123"), "JEFATURA"),
                "worker": create_user(db, UserCreate(username="worker-supervision", password="Clave-Worker-Supervision-123"), "TRABAJADOR"),
            }
            company = db.scalar(select(Empresa))
            self.workers = {
                "ana": Trabajador(usuario_id=self.users["worker"].id, empresa_id=company.id, codigo_interno="SUP-001", nombres="Ana", apellidos="Pérez"),
                "inactive": Trabajador(empresa_id=company.id, codigo_interno="SUP-002", nombres="Beto", apellidos="Histórico", activo=False),
                "empty": Trabajador(empresa_id=company.id, codigo_interno="SUP-003", nombres="Carla", apellidos="Sin Actividad"),
            }
            db.add_all(self.workers.values())
            db.flush()
            db.add(
                TarifaProvisionalAsistencia(
                    valor_clp=Decimal("30000"),
                    vigente_desde=date(2026, 9, 1),
                    origen="SISTEMA",
                )
            )
            db.commit()
            self.day_shift_id = db.scalar(select(Turno.id).where(Turno.codigo == "DIURNO"))
            self.night_shift_id = db.scalar(select(Turno.id).where(Turno.codigo == "NOCTURNO"))
            self.place_id = db.scalar(select(LugarTrabajo.id))
            self.worker_ids = {key: value.id for key, value in self.workers.items()}

            self.incomplete_id, self.pending_incident_id = self._add_session(
                db,
                self.workers["ana"].id,
                date(2026, 9, 2),
                self.day_shift_id,
                (9, 0),
                None,
                geofence="DENTRO_TOLERANCIA",
                incidents=("GPS_BAJA_PRECISION",),
            )
            self._add_session(db, self.workers["ana"].id, date(2026, 9, 3), self.day_shift_id, (9, 0), (18, 0), geofence="FUERA_RANGO", incidents=("FUERA_RANGO",))
            self._add_session(db, self.workers["ana"].id, date(2026, 9, 3), self.night_shift_id, (19, 0), (5, 0), exit_day=date(2026, 9, 4))
            self._add_session(db, self.workers["inactive"].id, date(2026, 9, 4), self.day_shift_id, (9, 0), (18, 0))
            self._add_session(db, self.workers["ana"].id, date(2026, 8, 31), self.day_shift_id, (9, 0), (18, 0))

        def override_db():
            with self.Session() as db:
                yield db

        app.dependency_overrides[get_db] = override_db

    def tearDown(self) -> None:
        app.dependency_overrides.clear()
        self.engine.dispose()
        for name, value in self.previous.items():
            if value is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = value

    @staticmethod
    def _local_utc(day: date, hour: int, minute: int) -> datetime:
        return datetime.combine(day, time(hour, minute), tzinfo=app_timezone()).astimezone(timezone.utc)

    def _add_session(
        self,
        db,
        worker_id,
        operational_day,
        shift_id,
        entry,
        exit_time,
        *,
        exit_day=None,
        geofence=None,
        incidents=(),
    ):
        session = SesionTrabajo(
            trabajador_id=worker_id,
            turno_id=shift_id,
            fecha_operacional=operational_day,
            estado="CERRADA" if exit_time else "ABIERTA",
            cerrado_at=self._local_utc(exit_day or operational_day, *exit_time) if exit_time else None,
        )
        db.add(session)
        db.flush()
        entry_mark = MarcajeAsistencia(sesion_id=session.id, tipo="ENTRADA", ocurrido_at=self._local_utc(operational_day, *entry))
        db.add(entry_mark)
        db.flush()
        if geofence:
            db.add(
                EvaluacionGeograficaMarcaje(
                    marcaje_id=entry_mark.id,
                    lugar_detectado_id=self.place_id,
                    distancia_m=Decimal("25"),
                    tolerancia_m_aplicada=Decimal("100"),
                    tipo_geocerca_aplicado="COMUNA",
                    geometria_version="subdere-test",
                    estado_geocerca=geofence,
                    estado_precision="BAJA_PRECISION" if "GPS_BAJA_PRECISION" in incidents else "ACEPTABLE",
                    max_precision_m_aplicada=Decimal("100"),
                    regla_version="test",
                    evaluada_at=self._local_utc(operational_day, *entry),
                )
            )
        incident_id = None
        for kind in incidents:
            incident = IncidenciaAsistencia(marcaje_id=entry_mark.id, tipo=kind)
            db.add(incident)
            db.flush()
            incident_id = incident_id or incident.id
        if exit_time:
            db.add(MarcajeAsistencia(sesion_id=session.id, tipo="SALIDA", ocurrido_at=self._local_utc(exit_day or operational_day, *exit_time)))
        db.commit()
        return session.id, incident_id

    @staticmethod
    def login(client, username, password):
        page = client.get("/login")
        token = re.search(r'name="csrf_token" value="([^"]+)"', page.text).group(1)
        return client.post("/login", {"username": username, "password": password, "csrf_token": token})

    def client_for(self, role):
        client = ASGIClient()
        credentials = {
            "admin": ("admin-supervision", "Clave-Admin-Supervision-123"),
            "jefatura": ("jefatura-supervision", "Clave-Jefatura-Supervision-123"),
            "worker": ("worker-supervision", "Clave-Worker-Supervision-123"),
        }[role]
        self.assertEqual(self.login(client, *credentials).status_code, 303)
        return client

    def test_service_filters_period_search_history_and_summary(self):
        period = supervision_period("2026-09-01", "2026-09-30")
        with self.Session() as db:
            listing = list_worker_supervision(db, period)
            by_code = {item.worker.codigo_interno: item for item in listing.workers}
            self.assertEqual(set(by_code), {"SUP-001", "SUP-002", "SUP-003"})
            self.assertFalse(by_code["SUP-002"].worker.activo)
            self.assertEqual(by_code["SUP-003"].projection.activity_days, 0)
            ana = by_code["SUP-001"].projection
            self.assertEqual((ana.activity_days, ana.completed_worked_days), (2, 1))
            self.assertEqual((ana.payable_shifts, ana.double_shift_days), (3, 1))
            self.assertEqual((ana.incident_count, ana.provisional_total_clp), (2, 90000))
            searched = list_worker_supervision(db, period, query="SUP-002")
            self.assertEqual([item.worker.nombres for item in searched.workers], ["Beto"])
            searched_by_name = list_worker_supervision(db, period, query="Ana Pérez")
            self.assertEqual([item.worker.codigo_interno for item in searched_by_name.workers], ["SUP-001"])
            wildcard = list_worker_supervision(db, period, query="%")
            self.assertEqual(wildcard.total_workers, 0)

        with self.assertRaises(AttendanceSupervisionError):
            supervision_period("2026-09-30", "2026-09-01")
        with self.assertRaises(AttendanceSupervisionError):
            supervision_period("2025-01-01", "2026-09-30")

    def test_day_detail_covers_double_night_geofence_incidents_and_missing_rate(self):
        with self.Session() as db:
            result = get_worker_day_supervision(db, self.worker_ids["ana"], date(2026, 9, 3))
            _, day = result
            self.assertTrue(day.projection.is_double_shift)
            self.assertIn("DOBLE TURNO", day.labels)
            night = next(item for item in day.sessions if item.projection.shift_code == "NOCTURNO")
            self.assertEqual(night.projection.exit_time.strftime("%d/%m %H:%M"), "04/09 05:00")
            outside = next(mark for session in day.sessions for mark in session.marks if mark.geofence_status)
            self.assertEqual(outside.geofence_status, "FUERA_RANGO")
            self.assertEqual(outside.incidents[0].kind, "FUERA_RANGO")

            _, early = get_worker_day_supervision(db, self.worker_ids["ana"], date(2026, 8, 31))
            self.assertIsNone(early.projection.effective_rate)
            self.assertIsNone(early.projection.provisional_total_clp)

    def test_get_routes_rbac_navigation_and_privacy(self):
        paths = (
            "/asistencia/supervision?fecha_desde=2026-09-01&fecha_hasta=2026-09-30",
            f"/asistencia/supervision/trabajadores/{self.worker_ids['ana']}?fecha_desde=2026-09-01&fecha_hasta=2026-09-30",
            f"/asistencia/supervision/trabajadores/{self.worker_ids['ana']}/dias/2026-09-02",
        )
        self.assertEqual(ASGIClient().get(paths[0]).status_code, 303)
        worker = self.client_for("worker")
        self.assertEqual(worker.get(paths[0]).status_code, 403)
        worker_dashboard = worker.get("/mi-asistencia")
        self.assertNotIn("Supervisión asistencia", unescape(worker_dashboard.text))
        for role in ("admin", "jefatura"):
            client = self.client_for(role)
            for path in paths:
                response = client.get(path)
                self.assertEqual(response.status_code, 200)
                self.assertNotIn("-33.2", response.text)
                self.assertNotIn("-70.6", response.text)
            self.assertIn("Supervisión asistencia", unescape(client.get(paths[0]).text))
        detail = self.client_for("admin").get(paths[2])
        self.assertIn("Jornada incompleta: falta SALIDA", detail.text)
        self.assertIn("DENTRO_TOLERANCIA", detail.text)
        self.assertIn("GPS_BAJA_PRECISION", detail.text)

    def test_administrative_exit_http_success_validation_csrf_rbac_and_conflict(self):
        path = f"/asistencia/supervision/sesiones/{self.incomplete_id}/completar-salida"
        payload = {"salida_at": "2026-09-02T18:00", "motivo": "Olvido confirmado."}
        admin = self.client_for("admin")
        self.assertEqual(ASGIClient().post(path, payload).status_code, 303)
        self.assertEqual(admin.post(path, payload).status_code, 403)
        invalid = dict(payload, csrf_token="incorrecto")
        self.assertEqual(admin.post(path, invalid).status_code, 403)
        worker = self.client_for("worker")
        self.assertEqual(worker.post(path, dict(payload, csrf_token=worker.cookies["boliklor_csrf"])).status_code, 403)

        too_early = admin.post(path, dict(payload, salida_at="2026-09-02T09:04", csrf_token=admin.cookies["boliklor_csrf"]))
        self.assertEqual(too_early.status_code, 409)
        empty_reason = admin.post(path, dict(payload, motivo=" ", csrf_token=admin.cookies["boliklor_csrf"]))
        self.assertEqual(empty_reason.status_code, 422)
        invalid_datetime = admin.post(path, dict(payload, salida_at="no-es-fecha", csrf_token=admin.cookies["boliklor_csrf"]))
        self.assertEqual(invalid_datetime.status_code, 422)
        success = admin.post(path, dict(payload, csrf_token=admin.cookies["boliklor_csrf"]))
        self.assertEqual(success.status_code, 303)
        refreshed = admin.get(success.headers["location"])
        self.assertIn("SALIDA administrativa registrada correctamente.", refreshed.text)
        self.assertIn("SALIDA agregada administrativamente", refreshed.text)
        self.assertNotIn("Jornada incompleta: falta SALIDA", refreshed.text)
        conflict = admin.post(path, dict(payload, csrf_token=admin.cookies["boliklor_csrf"]))
        self.assertEqual(conflict.status_code, 409)
        self.assertIn("La sesión ya fue completada por otro usuario.", conflict.text)

    def test_incident_http_decisions_csrf_rbac_comment_and_final_state(self):
        path = f"/asistencia/supervision/incidencias/{self.pending_incident_id}/decision"
        admin = self.client_for("admin")
        self.assertEqual(admin.post(path, {"decision": "APROBADA"}).status_code, 403)
        self.assertEqual(admin.post(path, {"decision": "APROBADA", "csrf_token": "incorrecto"}).status_code, 403)
        anonymous = ASGIClient()
        self.assertEqual(anonymous.post(path, {"decision": "APROBADA"}).status_code, 303)
        worker = self.client_for("worker")
        self.assertEqual(worker.post(path, {"decision": "APROBADA", "csrf_token": worker.cookies["boliklor_csrf"]}).status_code, 403)
        invalid_decision = admin.post(path, {"decision": "RESUELTA", "csrf_token": admin.cookies["boliklor_csrf"]})
        self.assertEqual(invalid_decision.status_code, 422)

        success = admin.post(path, {"decision": "RECHAZADA", "comentario": "Evidencia insuficiente.", "csrf_token": admin.cookies["boliklor_csrf"]})
        self.assertEqual(success.status_code, 303)
        page = admin.get(success.headers["location"])
        self.assertIn("RECHAZADA", page.text)
        self.assertIn("Evidencia insuficiente.", page.text)
        self.assertNotIn('value="APROBADA"', page.text)
        with self.Session() as db:
            stored = db.get(IncidenciaAsistencia, self.pending_incident_id)
            self.assertEqual((stored.estado, stored.comentario_resolucion), ("RECHAZADA", "Evidencia insuficiente."))
        conflict = admin.post(path, {"decision": "APROBADA", "csrf_token": admin.cookies["boliklor_csrf"]})
        self.assertEqual(conflict.status_code, 409)
        self.assertIn("La incidencia ya fue resuelta por otro usuario.", conflict.text)

    def test_jefatura_can_complete_exit_and_approve_incident(self):
        client = self.client_for("jefatura")
        csrf = client.cookies["boliklor_csrf"]
        completed = client.post(
            f"/asistencia/supervision/sesiones/{self.incomplete_id}/completar-salida",
            {
                "csrf_token": csrf,
                "salida_at": "2026-09-02T18:00",
                "motivo": "Revisión de Jefatura.",
            },
        )
        self.assertEqual(completed.status_code, 303)
        decided = client.post(
            f"/asistencia/supervision/incidencias/{self.pending_incident_id}/decision",
            {
                "csrf_token": csrf,
                "decision": "APROBADA",
                "comentario": "Aprobada por Jefatura.",
            },
        )
        self.assertEqual(decided.status_code, 303)
        with self.Session() as db:
            incident = db.get(IncidenciaAsistencia, self.pending_incident_id)
            self.assertEqual(incident.estado, "APROBADA")

    def test_invalid_period_and_unknown_targets_are_controlled(self):
        client = self.client_for("jefatura")
        invalid = client.get("/asistencia/supervision?fecha_desde=2026-09-30&fecha_hasta=2026-09-01")
        self.assertEqual(invalid.status_code, 422)
        self.assertIn("no puede ser posterior", invalid.text)
        self.assertEqual(client.get("/asistencia/supervision/trabajadores/999999").status_code, 404)
        csrf = client.cookies["boliklor_csrf"]
        self.assertEqual(client.post("/asistencia/supervision/sesiones/999999/completar-salida", {"csrf_token": csrf, "salida_at": "2026-09-02T18:00", "motivo": "x"}).status_code, 404)
        self.assertEqual(client.post("/asistencia/supervision/incidencias/999999/decision", {"csrf_token": csrf, "decision": "APROBADA"}).status_code, 404)

    def test_combined_excel_is_valid_filtered_safe_private_and_matches_projection(self):
        with self.Session() as db:
            company = db.scalar(select(Empresa))
            db.add_all(
                [
                    Trabajador(
                        empresa_id=company.id,
                        codigo_interno="@FORMULA",
                        nombres="=SUM(1,1)",
                        apellidos="Especial Ñ",
                    ),
                    *(
                        Trabajador(
                            empresa_id=company.id,
                            codigo_interno=f"LOTE-{index:02d}",
                            nombres="Carga",
                            apellidos=f"Zeta {index:02d}",
                        )
                        for index in range(27)
                    ),
                ]
            )
            db.commit()

        client = self.client_for("jefatura")
        path = "/asistencia/supervision/exportar?fecha_desde=2026-09-01&fecha_hasta=2026-09-30"
        response = client.get(path)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertRegex(
            response.headers["content-disposition"],
            r'^attachment; filename="asistencia_2026-09-01_2026-09-30\.xlsx"$',
        )
        workbook = load_workbook(BytesIO(response.body), read_only=True, data_only=False)
        sheet = workbook["Resumen"]
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            (
                "Trabajador",
                "Código",
                "Días trabajados",
                "Jornadas pagables",
                "Dobles turnos",
                "Incidencias",
                "Total provisional",
            ),
        )
        self.assertEqual(len(rows) - 1, 31)
        by_code = {row[1]: row for row in rows[1:]}
        ana_row = by_code["SUP-001"]
        with self.Session() as db:
            listing = list_worker_supervision(
                db, supervision_period("2026-09-01", "2026-09-30")
            )
            ana = next(
                item.projection
                for item in listing.workers
                if item.worker.codigo_interno == "SUP-001"
            )
        self.assertEqual(
            ana_row[2:],
            (
                ana.completed_worked_days,
                ana.payable_shifts,
                ana.double_shift_days,
                ana.incident_count,
                ana.provisional_total_clp,
            ),
        )
        formula_row = next(row for row in rows if row[1] == "'@FORMULA")
        self.assertEqual(formula_row[0], "'=SUM(1,1) Especial Ñ")
        serialized = " ".join(str(value) for row in rows for value in row if value is not None).casefold()
        self.assertNotIn("latitud", serialized)
        self.assertNotIn("longitud", serialized)
        self.assertNotIn("-33.2", serialized)
        self.assertNotIn("-70.6", serialized)

        filtered = client.get(f"{path}&q=SUP-002")
        filtered_rows = list(
            load_workbook(BytesIO(filtered.body), read_only=True)["Resumen"].iter_rows(
                values_only=True
            )
        )
        self.assertEqual(len(filtered_rows), 2)
        self.assertEqual(filtered_rows[1][1], "SUP-002")

    def test_individual_excel_has_summary_daily_detail_and_domain_semantics(self):
        path = (
            f"/asistencia/supervision/trabajadores/{self.worker_ids['ana']}/exportar"
            "?fecha_desde=2026-09-01&fecha_hasta=2026-09-30"
        )
        response = self.client_for("admin").get(path)
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.body), read_only=True, data_only=False)
        self.assertEqual(workbook.sheetnames, ["Resumen", "Detalle diario"])
        summary = dict(workbook["Resumen"].iter_rows(values_only=True))
        self.assertEqual(
            summary["Trabajador"],
            f"{self.workers['ana'].nombres} {self.workers['ana'].apellidos}",
        )
        self.assertEqual(summary["Período"], "2026-09-01 al 2026-09-30")
        self.assertEqual(
            tuple(summary[key] for key in ("Días trabajados", "Jornadas pagables", "Dobles turnos", "Incidencias", "Total provisional")),
            (1, 3, 1, 2, 90000),
        )
        detail_rows = list(workbook["Detalle diario"].iter_rows(values_only=True))
        headers = detail_rows[0]
        by_day = {row[0].date(): dict(zip(headers, row)) for row in detail_rows[1:]}
        incomplete = by_day[date(2026, 9, 2)]
        self.assertIn("INCOMPLETA", incomplete["Estado"])
        self.assertEqual(incomplete["Jornada pagable"], 1)
        self.assertEqual(incomplete["Incidencias"], 1)
        double = by_day[date(2026, 9, 3)]
        self.assertEqual(double["Turno"], "DIURNO + NOCTURNO")
        self.assertEqual(double["Jornada pagable"], 2)
        self.assertEqual(double["Tarifa efectiva"], 30000)
        self.assertEqual(double["Origen tarifa"], "GLOBAL")
        self.assertEqual(double["Total provisional"], 60000)
        self.assertIn("NOCTURNO: 04/09/2026 05:00", double["Salida"])

        early = self.client_for("jefatura").get(
            f"/asistencia/supervision/trabajadores/{self.worker_ids['ana']}/exportar"
            "?fecha_desde=2026-08-31&fecha_hasta=2026-08-31"
        )
        early_book = load_workbook(BytesIO(early.body), read_only=True)
        early_summary = dict(early_book["Resumen"].iter_rows(values_only=True))
        early_detail = list(early_book["Detalle diario"].iter_rows(values_only=True))
        self.assertEqual(early_summary["Total provisional"], "Sin tarifa configurada")
        self.assertEqual(early_detail[1][6], "Sin tarifa configurada")
        self.assertEqual(early_detail[1][8], "Sin tarifa configurada")

    def test_excel_routes_enforce_supervision_permission(self):
        paths = (
            "/asistencia/supervision/exportar?fecha_desde=2026-09-01&fecha_hasta=2026-09-30",
            f"/asistencia/supervision/trabajadores/{self.worker_ids['ana']}/exportar?fecha_desde=2026-09-01&fecha_hasta=2026-09-30",
        )
        for path in paths:
            self.assertEqual(ASGIClient().get(path).status_code, 303)
            self.assertEqual(self.client_for("worker").get(path).status_code, 403)
            self.assertEqual(self.client_for("admin").get(path).status_code, 200)
            self.assertEqual(self.client_for("jefatura").get(path).status_code, 200)
        self.assertEqual(
            self.client_for("admin").get(
                "/asistencia/supervision/exportar?fecha_desde=2026-09-30&fecha_hasta=2026-09-01"
            ).status_code,
            422,
        )
        self.assertEqual(
            self.client_for("admin").get(
                "/asistencia/supervision/trabajadores/999999/exportar?fecha_desde=2026-09-01&fecha_hasta=2026-09-30"
            ).status_code,
            404,
        )


if __name__ == "__main__":
    unittest.main()
