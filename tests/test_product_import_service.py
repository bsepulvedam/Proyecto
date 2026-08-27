import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from app.database.base import Base
from app.models.empresa import Empresa
from app.models.producto import Producto
from app.models.unidad_medida import UnidadMedida
from app.services.product_import_service import (
    ProductImportReport,
    ProductoImportRow,
    analyze_product_workbook,
    company_from_sku,
    import_valid_products,
    normalize_sku,
)


class ProductImportServiceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temporary.close()
        cls.workbook_path = Path(temporary.name)

        workbook = Workbook()
        master = workbook.active
        master.title = "Maestro de materiales"
        master.append(["SKU", "item", "unidad", "tipo", "familia", "unidad"])
        master.append([" bol-8 ", "PINTURA TERMOPLASTICA", "SACO", "MATERIAL", "DEMARCACION", 25])
        master.append(["BOL-20", "PRODUCTO FACTOR VEINTE", "SACO", "MATERIAL", "DEMARCACION", 20])
        master.append(["ALM-40", "KIT DE PRUEBA", "KIT", "MATERIAL", "INSTALACION", 1])
        master.append(["BOL-UNK", "UNIDAD DESCONOCIDA", "BALDE", "MATERIAL", "DEMARCACION", 1])
        master.append(["BOL-DUP", "DUPLICADO UNO", "UNIDAD", "MATERIAL", "DEMARCACION", 1])
        master.append(["BOL-DUP", "DUPLICADO DOS", "UNIDAD", "MATERIAL", "DEMARCACION", 1])
        master.append([None, "SIN SKU", "UNIDAD", "MATERIAL", "DEMARCACION", 1])
        master.append(["BOL-BAD", "FACTOR INVALIDO", "UNIDAD", "MATERIAL", "DEMARCACION", 0])

        boliklor = workbook.create_sheet("Stock Boliklor")
        for _ in range(3):
            boliklor.append([])
        boliklor.append(["SKU", "item", "unidad", "tipo", "familia", "Stock min"])
        boliklor.append(["BOL-8", "PINTURA TERMOPLASTICA", "SACO", "MATERIAL", "DEMARCACION", 80])
        boliklor.append(["BOL-20", "PRODUCTO FACTOR VEINTE", "SACO", "MATERIAL", "DEMARCACION", 20])
        boliklor.append(["BOL-UNK", "UNIDAD DESCONOCIDA", "BALDE", "MATERIAL", "DEMARCACION", 2])
        boliklor.append(["BOL-DUP", "DUPLICADO UNO", "UNIDAD", "MATERIAL", "DEMARCACION", 1])
        boliklor.append(["BOL-BAD", "FACTOR INVALIDO", "UNIDAD", "MATERIAL", "DEMARCACION", 0])

        alm = workbook.create_sheet("Stock ALM")
        for _ in range(3):
            alm.append([])
        alm.append(["SKU", "item", "unidad", "tipo", "familia", "Stock min"])
        alm.append(["ALM-40", "KIT DE PRUEBA", "KIT", "MATERIAL", "INSTALACION", 5])
        workbook.save(cls.workbook_path)
        workbook.close()

    @classmethod
    def tearDownClass(cls) -> None:
        cls.workbook_path.unlink(missing_ok=True)

    def setUp(self) -> None:
        companies = [SimpleNamespace(codigo="BOLIKLOR"), SimpleNamespace(codigo="ALM")]
        units = [
            SimpleNamespace(codigo=code)
            for code in ("UN", "KG", "SACO", "KIT")
        ]
        self.existing_products: list = []
        self.report = analyze_product_workbook(
            self.workbook_path, companies, units, self.existing_products
        )
        self.rows = {row.sku: row for row in self.report.rows if row.sku}

    def test_sku_normalization_and_company_mapping(self) -> None:
        self.assertEqual(normalize_sku(" bol - 8 "), "BOL-8")
        self.assertEqual(company_from_sku("BOL-8"), "BOLIKLOR")
        self.assertEqual(company_from_sku("ALM-40"), "ALM")

    def test_factor_25_and_factor_20_map_sack_content_to_kg(self) -> None:
        for sku, factor in (("BOL-8", 25), ("BOL-20", 20)):
            row = self.rows[sku]
            self.assertEqual(row.unidad_stock, "SACO")
            self.assertEqual(row.unidad_contenido, "KG")
            self.assertEqual(row.unidad_costo, "KG")
            self.assertEqual(row.factor_conversion, factor)

    def test_kit_remains_an_indivisible_stock_and_cost_unit(self) -> None:
        row = self.rows["ALM-40"]
        self.assertEqual(row.empresa, "ALM")
        self.assertEqual(row.unidad_stock, "KIT")
        self.assertIsNone(row.unidad_contenido)
        self.assertEqual(row.unidad_costo, "KIT")

    def test_unknown_unit_is_reported(self) -> None:
        row = self.rows["BOL-UNK"]
        self.assertEqual(row.validation_status, "ERROR")
        self.assertIn("BALDE", self.report.unidades_no_registradas)
        self.assertTrue(any("no registrada" in error for error in row.errores))

    def test_duplicate_sku_is_reported_on_each_row(self) -> None:
        duplicate_rows = [row for row in self.report.rows if row.sku == "BOL-DUP"]
        self.assertEqual(len(duplicate_rows), 2)
        self.assertIn("BOL-DUP", self.report.duplicados)
        self.assertTrue(all(row.validation_status == "ERROR" for row in duplicate_rows))

    def test_official_minimum_stock_comes_from_stock_sheet(self) -> None:
        self.assertEqual(self.rows["BOL-8"].stock_minimo, 80)
        self.assertEqual(self.rows["ALM-40"].stock_minimo, 5)

    def test_invalid_rows_are_not_ready_for_import(self) -> None:
        missing_sku = next(row for row in self.report.rows if not row.sku)
        invalid_factor = self.rows["BOL-BAD"]
        self.assertFalse(missing_sku.listo_para_importar)
        self.assertFalse(invalid_factor.listo_para_importar)
        self.assertTrue(any("mayor que 0" in error for error in invalid_factor.errores))

    def test_analysis_does_not_insert_or_mutate_existing_products(self) -> None:
        self.assertEqual(self.existing_products, [])
        self.assertEqual(len(self.report.rows), 8)


class ProductImportPersistenceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_engine("sqlite+pysqlite:///:memory:")
        Base.metadata.create_all(
            self.engine,
            tables=[Empresa.__table__, UnidadMedida.__table__, Producto.__table__],
        )
        self.db = Session(self.engine)
        self.db.add_all(
            [
                Empresa(codigo="BOLIKLOR", nombre="BOLIKLOR"),
                Empresa(codigo="ALM", nombre="ALM"),
                UnidadMedida(codigo="UN", nombre="Unidad", permite_decimales=False),
                UnidadMedida(codigo="KG", nombre="Kilogramo", permite_decimales=True),
                UnidadMedida(codigo="SACO", nombre="Saco", permite_decimales=False),
                UnidadMedida(codigo="KIT", nombre="Kit", permite_decimales=False),
            ]
        )
        self.db.commit()

    def tearDown(self) -> None:
        self.db.close()
        self.engine.dispose()

    @staticmethod
    def row(sku: str, company: str = "BOLIKLOR") -> ProductoImportRow:
        return ProductoImportRow(
            fila_excel=2,
            sku=sku,
            empresa=company,
            nombre="PRODUCTO CONTROLADO",
            descripcion=None,
            unidad_stock="SACO",
            unidad_contenido="KG",
            factor_conversion=Decimal("25"),
            unidad_costo="KG",
            stock_minimo=Decimal("80"),
            tipo="MATERIAL",
            familia="DEMARCACION",
        )

    @staticmethod
    def report(rows: list[ProductoImportRow]) -> ProductImportReport:
        return ProductImportReport(
            source_file="test.xlsx",
            sheets_used=["Maestro de materiales", "Stock Boliklor", "Stock ALM"],
            rows=rows,
            unidades_detectadas=["KG", "SACO"],
            unidades_no_registradas=[],
            duplicados=[],
            conflictos_maestro_stock=[],
            productos_stock_sin_maestro=[],
        )

    def test_only_fully_valid_rows_are_imported_with_exact_values(self) -> None:
        valid = self.row("BOL-OK")
        warning = self.row("ALM-WARN", "ALM")
        warning.advertencias.append("Revisión pendiente.")
        invalid = self.row("BOL-ERROR")
        invalid.errores.append("Unidad inconsistente.")

        result = import_valid_products(self.db, self.report([valid, warning, invalid]))
        products = list(self.db.scalars(select(Producto)).all())
        self.assertEqual(result.importados, 1)
        self.assertEqual(result.advertencias_pendientes, 1)
        self.assertEqual(result.errores_pendientes, 1)
        self.assertEqual(len(products), 1)
        product = products[0]
        self.assertEqual(product.sku, "BOL-OK")
        self.assertEqual(product.empresa.codigo, "BOLIKLOR")
        self.assertEqual(product.factor_conversion, Decimal("25.0000"))
        self.assertEqual(product.stock_minimo, Decimal("80.000"))
        self.assertEqual(product.unidad_stock.codigo, "SACO")
        self.assertEqual(product.unidad_contenido.codigo, "KG")
        self.assertEqual(product.unidad_costo.codigo, "KG")

    def test_existing_sku_is_idempotent_and_not_overwritten(self) -> None:
        report = self.report([self.row("BOL-IDEMPOTENT")])
        first = import_valid_products(self.db, report)
        second = import_valid_products(self.db, report)
        count = self.db.scalar(select(func.count()).select_from(Producto))
        self.assertEqual(first.importados, 1)
        self.assertEqual(second.importados, 0)
        self.assertEqual(second.existentes, 1)
        self.assertEqual(count, 1)

    def test_unresolvable_unit_is_not_created_or_imported(self) -> None:
        unresolved = self.row("BOL-UNKNOWN")
        unresolved.unidad_stock = "BALDE"
        result = import_valid_products(self.db, self.report([unresolved]))
        count = self.db.scalar(select(func.count()).select_from(Producto))
        unit_count = self.db.scalar(select(func.count()).select_from(UnidadMedida))
        self.assertEqual(result.importados, 0)
        self.assertEqual(result.total_fallidos, 1)
        self.assertEqual(count, 0)
        self.assertEqual(unit_count, 4)


if __name__ == "__main__":
    unittest.main()
