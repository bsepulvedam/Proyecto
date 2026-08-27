import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

from app.services.product_import_correction_service import analyze_product_corrections


def unit(code):
    return SimpleNamespace(codigo=code)


def product(sku, name):
    return SimpleNamespace(sku=sku, nombre=name)


class ProductImportCorrectionTests(unittest.TestCase):
    def setUp(self):
        temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temporary.close()
        self.path = Path(temporary.name)
        workbook = Workbook()
        master = workbook.active
        master.title = "Maestro de materiales"
        master.append(["SKU", "item", "unidad", "tipo", "familia", "unidad"])
        master_rows = [
            ["BOL-1", "PINTURA ACRILICA AMARILLA", "tineta", "MATERIAL", "DEMARCACION", 1],
            ["BOL-8", "PINTURA TERMOPLASTICA BLANCA", "saco", "MATERIAL", "DEMARCACION", 25],
            ["BOL-56", "TIZA EN POLVO", "kg", "MATERIAL", "DEMARCACION", 1],
            ["BOL-10", "PINTURA TERMOPLASTICA AMARILLA", "SACO", "MATERIAL", "DEMARCACION", 25],
        ]
        for row in master_rows:
            master.append(row)

        stock = workbook.create_sheet("Stock Boliklor")
        for _ in range(3):
            stock.append([])
        stock.append(["SKU", "item", "unidad", "tipo", "familia", "Stock min"])
        for row, minimum in zip(master_rows, (10, 80, 10, 80)):
            stock.append([*row[:5], minimum])
        stock_alm = workbook.create_sheet("Stock ALM")
        for _ in range(3):
            stock_alm.append([])
        stock_alm.append(["SKU", "item", "unidad", "tipo", "familia", "Stock min"])

        reception = workbook.create_sheet("Recepción")
        reception.append(["codigo", "Descripción", "fecha", "Cantidad recibida", "Costo unitario", "unidad", "$ precio de compra", "$ Costo promedio", "Stock Empresa", "unidad"])
        reception.append(["BOL-1", "PINTURA", None, 47, 42300, 1, 1988100, None, "Boliklor", "TINETA"])
        reception.append(["BOL-8", "TERMOPLASTICA", None, 280, 880, 25, 6160000, None, "Boliklor", "SACO"])
        reception.append(["BOL-10", "TERMOPLASTICA", None, 202, 880, 25, 4444000, None, "Boliklor", "SACO"])
        dispatch = workbook.create_sheet("Despacho")
        dispatch.append(["SKU", None, None, None, "Producto", "Salida"])
        dispatch.append(["BOL-8", None, None, None, "TERMOPLASTICA", 275])
        dispatch.append(["BOL-10", None, None, None, "TERMOPLASTICA", 188])
        returns = workbook.create_sheet("Devoluciones")
        returns.append(["SKU", "Producto", "fecha", "Cantidad recibida"])
        returns.append(["BOL-8", "TERMOPLASTICA", None, 31])
        returns.append(["BOL-10", "TERMOPLASTICA", None, 65])
        transformation = workbook.create_sheet("transformacion de unidad")
        transformation.append(["SKU", "factor"])
        for sku in ("BOL-1", "BOL-8", "BOL-56", "BOL-10"):
            transformation.append([sku, 1000])
        workbook.save(self.path)
        workbook.close()
        self.products = [product(row[0], row[1]) for row in master_rows]
        self.units = [unit(code) for code in ("UN", "KG", "SACO", "TINETA")]

    def tearDown(self):
        self.path.unlink(missing_ok=True)

    def report(self):
        return analyze_product_corrections(self.path, self.products, self.units)

    def test_required_product_examples(self):
        rows = {row.sku: row for row in self.report().rows}
        bol1 = rows["BOL-1"]
        self.assertEqual((bol1.unidad_presentacion, bol1.stock_actual, bol1.stock_minimo, bol1.factor_conversion, bol1.unidad_costo, bol1.estado_stock), ("TINETA", 47, 10, 1, "TINETA", "EN STOCK"))
        bol8 = rows["BOL-8"]
        self.assertEqual((bol8.unidad_presentacion, bol8.stock_actual, bol8.stock_minimo, bol8.factor_conversion, bol8.unidad_contenido, bol8.unidad_costo, bol8.estado_stock), ("SACO", 36, 80, 25, "KG", "KG", "EN STOCK"))
        bol56 = rows["BOL-56"]
        self.assertEqual((bol56.unidad_presentacion, bol56.stock_actual, bol56.stock_minimo, bol56.factor_conversion, bol56.unidad_contenido, bol56.unidad_costo, bol56.estado_stock), ("KG", 0, 10, 1, None, "KG", "SIN STOCK"))
        self.assertEqual(rows["BOL-10"].stock_actual, 79)

    def test_quantities_render_without_artificial_decimals(self):
        rows = {row.sku: row for row in self.report().rows}
        self.assertEqual(rows["BOL-1"].stock_minimo_display, "10")
        self.assertEqual(rows["BOL-8"].stock_actual_display, "36")
        self.assertEqual(rows["BOL-8"].factor_display, "25")

    def test_transformation_sheet_does_not_determine_factor(self):
        factors = {row.factor_conversion for row in self.report().rows}
        self.assertEqual(factors, {Decimal("1"), Decimal("25")})
        self.assertNotIn(Decimal("1000"), factors)

    def test_preview_does_not_mutate_products(self):
        before = [(p.sku, p.nombre) for p in self.products]
        self.report()
        self.assertEqual(before, [(p.sku, p.nombre) for p in self.products])


if __name__ == "__main__":
    unittest.main()
